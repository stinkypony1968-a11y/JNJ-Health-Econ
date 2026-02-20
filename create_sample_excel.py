#!/usr/bin/env python3
"""Create sample_hospitals.xlsx with the v2 health economics columns."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Hospitals"

headers = [
    "Hospital Name", "Short Name", "Address", "City", "State", "County",
    "Latitude", "Longitude", "Phone", "Licensed Beds", "Affiliation",
    "Buy Group/GPO", "Stroke Certification", "Strokes Per Year",
    "IR Phone", "Cert Body", "Neurosurgery Fellowship", "Neuro IR Fellowship",
    "Catchment Population", "Population 65+", "Median Age", "Life Expectancy",
    "CAH Status", "Telestroke Capable", "24/7 Thrombectomy",
    "tPA Available", "Neuro ICU", "CT Scanner", "Spoke Hospital",
    "A-Fib Prevalence", "A-Fib Absolute Count", "MMAE Score",
    "Clinical Benefit Radius (km)",
    "Medevac Available", "Road Access", "Comment"
]

header_font = Font(bold=True, size=10)
header_fill = PatternFill(start_color="00205B", end_color="00205B", fill_type="solid")
header_font_white = Font(bold=True, size=10, color="FFFFFF")

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

data = [
    ["Baptist Medical Center Jacksonville", "Baptist Jax", "800 Prudential Dr, Jacksonville, FL 32207", "Jacksonville", "FL", "Duval County, FL",
     30.3187, -81.6568, "(904) 202-2000", 972, "Baptist Health", "Vizient", "CSC", 900,
     "(904) 202-2100", "Joint Commission", True, True,
     1750000, 310000, 37.5, 77.4,
     False, True, True, True, True, True, False,
     3.2, 56000, 72, 45,
     True, True, "Flagship CSC — 175% YoY growth in thrombectomy"],

    ["UF Health Jacksonville", "UF Health Jax", "655 W 8th St, Jacksonville, FL 32209", "Jacksonville", "FL", "Duval County, FL",
     30.3554, -81.6721, "(904) 244-0411", 695, "University of Florida Health", "Vizient", "CSC", 750,
     "(904) 244-0500", "Joint Commission", True, True,
     1400000, 238000, 36.8, 76.9,
     False, True, True, True, True, True, False,
     2.9, 40600, 68, 40,
     True, True, "Academic medical center with active NIR fellowship"],

    ["Mayo Clinic Jacksonville", "Mayo Jax", "4500 San Pablo Rd S, Jacksonville, FL 32224", "Jacksonville", "FL", "Duval County, FL",
     30.2617, -81.4408, "(904) 953-2000", 304, "Mayo Clinic Health System", "Mayo GPO", "CSC", 420,
     "(904) 953-2100", "Joint Commission", True, True,
     900000, 171000, 42.1, 81.2,
     False, True, True, True, True, True, False,
     3.8, 34200, 85, 50,
     True, True, "Research-intensive — leading DEFUSE trials enrollment"],

    ["Ascension St. Vincent's Riverside", "St. Vincent's", "1 Shircliff Way, Jacksonville, FL 32204", "Jacksonville", "FL", "Duval County, FL",
     30.3136, -81.6892, "(904) 308-7300", 528, "Ascension Health", "HPG", "PSC", 340,
     "", "DNV", False, False,
     850000, 144500, 39.2, 78.1,
     False, True, False, True, True, True, False,
     2.5, 21250, 55, 30,
     False, True, "Growing stroke program under new medical director"],

    ["Memorial Hospital Jacksonville", "Memorial Jax", "3625 University Blvd S, Jacksonville, FL 32216", "Jacksonville", "FL", "Duval County, FL",
     30.2843, -81.6021, "(904) 702-6111", 453, "HCA Healthcare", "HealthTrust", "PSC", 280,
     "", "Joint Commission", False, False,
     700000, 119000, 38.4, 77.8,
     False, True, False, True, True, True, False,
     2.1, 14700, 48, 25,
     False, True, ""],

    ["Flagler Hospital", "Flagler", "400 Health Park Blvd, St. Augustine, FL 32086", "St. Augustine", "FL", "St. Johns County, FL",
     29.8723, -81.3168, "(904) 819-5155", 335, "Flagler Health+", "Premier", "PSC", 210,
     "", "Joint Commission", False, False,
     350000, 77000, 44.6, 79.5,
     False, True, False, True, False, True, True,
     3.1, 10850, 42, 20,
     False, True, "Spoke to Baptist Jax — telestroke hub model"],

    ["Orange Park Medical Center", "Orange Park", "2001 Kingsley Ave, Orange Park, FL 32073", "Orange Park", "FL", "Clay County, FL",
     30.1665, -81.7178, "(904) 639-8500", 317, "HCA Healthcare", "HealthTrust", "TSC", 180,
     "", "DNV", False, False,
     280000, 47600, 36.9, 77.2,
     False, True, False, True, False, True, True,
     1.8, 5040, 35, 15,
     False, True, ""],

    ["Southeast Georgia Health System - Brunswick", "SGHS Brunswick", "2415 Parkwood Dr, Brunswick, GA 31520", "Brunswick", "GA", "Glynn County, GA",
     31.1496, -81.4746, "(912) 466-7000", 532, "Southeast Georgia Health System", "Vizient", "PSC", 190,
     "", "Joint Commission", False, False,
     320000, 60800, 40.3, 75.8,
     False, True, False, True, True, True, False,
     2.4, 7680, 40, 25,
     True, True, "Nearest CSC transfer is 90 min — medevac critical"],

    ["Memorial Health University Medical Center", "Memorial Savannah", "4700 Waters Ave, Savannah, GA 31404", "Savannah", "GA", "Chatham County, GA",
     32.0312, -81.0874, "(912) 350-8000", 612, "HCA Healthcare", "HealthTrust", "CSC", 580,
     "(912) 350-8200", "Joint Commission", True, False,
     680000, 108800, 35.7, 75.2,
     False, True, True, True, True, True, False,
     2.7, 18360, 62, 35,
     True, True, "Regional CSC hub for southeast Georgia"],

    ["Putnam Community Medical Center", "Putnam", "611 Zeagler Dr, Palatka, FL 32177", "Palatka", "FL", "Putnam County, FL",
     29.6516, -81.6568, "(386) 328-5711", 99, "ScionHealth", "HealthTrust", "None", 45,
     "", "", False, False,
     75000, 15000, 42.8, 73.6,
     True, False, False, True, False, True, True,
     1.5, 1125, 18, 10,
     False, True, "Critical access — rural area, limited neurovascular capability"],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Set column widths
widths = {1: 42, 2: 20, 3: 40, 4: 16, 5: 6, 6: 22, 7: 10, 8: 10, 9: 16, 10: 12,
          11: 30, 12: 14, 13: 18, 14: 14, 15: 16, 16: 18, 17: 12, 18: 12,
          19: 16, 20: 14, 21: 10, 22: 12, 23: 10, 24: 12, 25: 14, 26: 10,
          27: 10, 28: 10, 29: 12, 30: 12, 31: 14, 32: 10, 33: 14, 34: 12, 35: 10, 36: 40}
for col, width in widths.items():
    ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = width

wb.save("/sessions/confident-beautiful-turing/jnj-territory-maps/sample_hospitals.xlsx")
print("Created sample_hospitals.xlsx with 10 hospitals and v2 health economics columns")

#!/usr/bin/env python3
"""
Territory Map Generator for J&J MedTech Neurovascular
=====================================================
Health Economics & Clinical Infrastructure Edition

Generates a single-file HTML territory map from an Excel file of hospital data.
Focused on stroke epidemiology, clinical infrastructure, and geographic access.
No sales/revenue data — purely clinical and health economics.

Usage:
    python generate_territory_map.py --excel hospitals.xlsx --config config.json --output territory-map.html
    python generate_territory_map.py hospitals.xlsx --rep "Andrew Payne" --territory "Florida / Georgia" --output flga-map.html
"""

import argparse
import json
import math
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from jinja2 import Template

# ─── Constants ───────────────────────────────────────────────────────────────

REQUIRED_COLUMNS = [
    "Hospital Name", "Short Name", "Address", "City", "State", "County",
    "Latitude", "Longitude", "Phone", "Licensed Beds", "Affiliation",
    "Buy Group/GPO", "Stroke Certification", "Strokes Per Year"
]

OPTIONAL_COLUMNS = [
    "IR Phone", "Cert Body", "Neurosurgery Fellowship", "Neuro IR Fellowship",
    "Catchment Population", "Population 65+", "Median Age", "Life Expectancy",
    "CAH Status", "Telestroke Capable", "24/7 Thrombectomy",
    "tPA Available", "Neuro ICU", "CT Scanner", "Spoke Hospital",
    "A-Fib Prevalence", "A-Fib Absolute Count", "MMAE Score",
    "Clinical Benefit Radius (km)",
    "Medevac Available", "Road Access",
    "Comment"
]

CERT_COLORS = {
    "CSC": "#EB1700", "PSC": "#F59E0B", "TSC": "#0077C8",
    "TCC": "#0077C8", "None": "#6B7280", "": "#6B7280"
}
CERT_LABELS = {
    "CSC": "Comprehensive Stroke Center", "PSC": "Primary Stroke Center",
    "TSC": "Thrombectomy-Capable Stroke Center", "TCC": "Thrombectomy-Capable Center",
    "None": "No Certification", "": "No Certification"
}

# Epidemiology rates (Rai et al. Stroke 2020, GBD 2019, AHA 2024, Rauhala et al. JAMA Neuro 2020)
EPIRATES = {
    "ischemic_per_100k": 216,
    "lvo_pct_of_ischemic": 0.21,
    "mt_eligibility_pct": 0.70,
    "avm_aneurysm_per_100k": 12,
    "hemorrhagic_per_100k": 12,
    "hemorrhagic_sah_pct": 0.05,
    "hemorrhagic_ich_pct": 0.80,
    "hemorrhagic_other_pct": 0.15,
    "age_65_multiplier": 2.5,
    "default_65_plus_pct": 0.17,
    "avg_los_ischemic_days": 6.6,
    "avg_charges_ischemic": 72787,
}

INFRA_WEIGHTS = {
    "cert_csc": 25, "cert_psc": 20, "cert_tsc": 15, "cert_none": 5,
    "thrombectomy_24_7": 15, "neuro_icu": 12, "cah": 10,
    "ct_scanner": 8, "telestroke": 8, "tpa": 5,
    "nsg_fellowship": 7, "nir_fellowship": 8, "stroke_volume_max": 10,
}


# ─── Data Loading ────────────────────────────────────────────────────────────

def load_excel(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        print(f"ERROR: Missing required columns: {', '.join(missing)}")
        sys.exit(1)
    hospitals = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        record = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        hospitals.append(record)
    wb.close()
    return hospitals


def load_config(config_path=None, args=None):
    config = {
        "rep_name": "Territory Rep", "rep_role": "CAS", "rep_base_city": "",
        "rep_base_lat": 0, "rep_base_lng": 0, "territory_name": "Territory",
        "map_center_lat": 0, "map_center_lng": 0, "map_zoom": 7, "team_members": []
    }
    if config_path and os.path.exists(config_path):
        with open(config_path, "r") as f:
            config.update(json.load(f))
    if args:
        if args.rep: config["rep_name"] = args.rep
        if args.territory: config["territory_name"] = args.territory
    return config


# ─── Helpers ─────────────────────────────────────────────────────────────────

def safe_float(val, default=0.0):
    if val is None: return default
    try: return float(val)
    except (ValueError, TypeError): return default

def safe_int(val, default=0):
    if val is None: return default
    try: return int(float(val))
    except (ValueError, TypeError): return default

def safe_bool(val, default=False):
    if val is None: return default
    if isinstance(val, bool): return val
    if isinstance(val, str): return val.lower() in ("true", "yes", "1", "y")
    return bool(val)


# ─── Data Enrichment ─────────────────────────────────────────────────────────

def enrich_hospital(h):
    # Basic fields
    h["name"] = str(h.get("Hospital Name", ""))
    h["short_name"] = str(h.get("Short Name", ""))
    h["address"] = str(h.get("Address", ""))
    h["city"] = str(h.get("City", ""))
    h["state"] = str(h.get("State", ""))
    h["county"] = str(h.get("County", ""))
    h["phone"] = str(h.get("Phone", ""))
    h["ir_phone"] = str(h.get("IR Phone", "")) if h.get("IR Phone") else ""
    h["lat"] = safe_float(h.get("Latitude"))
    h["lng"] = safe_float(h.get("Longitude"))
    h["beds"] = safe_int(h.get("Licensed Beds"))
    h["affiliation"] = str(h.get("Affiliation", ""))
    h["gpo"] = str(h.get("Buy Group/GPO", ""))
    h["strokes_yr"] = safe_int(h.get("Strokes Per Year"))
    h["comment"] = str(h.get("Comment", "")) if h.get("Comment") else ""

    # Certification
    h["cert"] = str(h.get("Stroke Certification", "None")).strip()
    if h["cert"] not in CERT_COLORS: h["cert"] = "None"
    h["cert_body"] = str(h.get("Cert Body", "")) if h.get("Cert Body") else ""
    h["cert_color"] = CERT_COLORS.get(h["cert"], "#6B7280")
    h["cert_label"] = CERT_LABELS.get(h["cert"], "No Certification")

    # Fellowships
    h["nsg_fellowship"] = safe_bool(h.get("Neurosurgery Fellowship"))
    h["nir_fellowship"] = safe_bool(h.get("Neuro IR Fellowship"))

    # Demographics
    h["median_age"] = safe_float(h.get("Median Age"))
    h["life_expectancy"] = safe_float(h.get("Life Expectancy"))
    pop_total = safe_int(h.get("Catchment Population"))
    pop_65 = safe_int(h.get("Population 65+"))
    if pop_total > 0 and pop_65 == 0:
        pop_65 = int(pop_total * EPIRATES["default_65_plus_pct"])
    h["pop_total"] = pop_total
    h["pop_65_plus"] = pop_65
    h["pop_under_65"] = max(0, pop_total - pop_65)
    h["pop_65_pct"] = round(pop_65 / pop_total * 100, 1) if pop_total > 0 else 0
    h["effective_pop"] = round(h["pop_under_65"] + pop_65 * EPIRATES["age_65_multiplier"]) if pop_total > 0 else 0

    # Stroke epidemiology
    if h["effective_pop"] > 0:
        ep = h["effective_pop"]
        h["ischemic_volume"] = round(ep * EPIRATES["ischemic_per_100k"] / 100000)
        h["lvo_volume"] = round(h["ischemic_volume"] * EPIRATES["lvo_pct_of_ischemic"])
        h["mt_eligible"] = round(h["lvo_volume"] * EPIRATES["mt_eligibility_pct"])
        h["avm_aneurysm_volume"] = round(ep * EPIRATES["avm_aneurysm_per_100k"] / 100000)
        h["hemorrhagic_volume"] = round(ep * EPIRATES["hemorrhagic_per_100k"] / 100000)
        h["sah_volume"] = round(h["hemorrhagic_volume"] * EPIRATES["hemorrhagic_sah_pct"])
        h["ich_volume"] = round(h["hemorrhagic_volume"] * EPIRATES["hemorrhagic_ich_pct"])
        h["hemorrhagic_other"] = round(h["hemorrhagic_volume"] * EPIRATES["hemorrhagic_other_pct"])
        h["total_stroke_volume"] = h["ischemic_volume"] + h["hemorrhagic_volume"]
    else:
        for k in ["ischemic_volume", "lvo_volume", "mt_eligible", "avm_aneurysm_volume",
                   "hemorrhagic_volume", "sah_volume", "ich_volume", "hemorrhagic_other", "total_stroke_volume"]:
            h[k] = 0

    # Infrastructure flags
    h["cah"] = safe_bool(h.get("CAH Status"))
    h["telestroke"] = safe_bool(h.get("Telestroke Capable"))
    h["thrombectomy_24_7"] = safe_bool(h.get("24/7 Thrombectomy"))
    h["tpa_available"] = safe_bool(h.get("tPA Available"))
    h["neuro_icu"] = safe_bool(h.get("Neuro ICU"))
    h["ct_scanner"] = safe_bool(h.get("CT Scanner"))
    h["spoke_hospital"] = safe_bool(h.get("Spoke Hospital"))

    # A-Fib (optional)
    h["afib_prevalence"] = safe_float(h.get("A-Fib Prevalence"))
    h["afib_count"] = safe_int(h.get("A-Fib Absolute Count"))
    h["mmae_score"] = safe_float(h.get("MMAE Score"))
    h["clinical_benefit_radius"] = safe_float(h.get("Clinical Benefit Radius (km)"))

    # Geographic access
    h["medevac_available"] = safe_bool(h.get("Medevac Available"))
    h["road_access"] = safe_bool(h.get("Road Access"), default=True)

    return h


# ─── Scoring & Tiers ────────────────────────────────────────────────────────

def calculate_infrastructure_scores(hospitals):
    max_strokes = max((h["strokes_yr"] for h in hospitals), default=1) or 1
    for h in hospitals:
        score = 0
        cert_map = {"CSC": INFRA_WEIGHTS["cert_csc"], "PSC": INFRA_WEIGHTS["cert_psc"],
                     "TSC": INFRA_WEIGHTS["cert_tsc"], "TCC": INFRA_WEIGHTS["cert_tsc"]}
        score += cert_map.get(h["cert"], INFRA_WEIGHTS["cert_none"])
        if h["thrombectomy_24_7"]: score += INFRA_WEIGHTS["thrombectomy_24_7"]
        if h["neuro_icu"]: score += INFRA_WEIGHTS["neuro_icu"]
        if h["cah"]: score += INFRA_WEIGHTS["cah"]
        if h["ct_scanner"]: score += INFRA_WEIGHTS["ct_scanner"]
        if h["telestroke"]: score += INFRA_WEIGHTS["telestroke"]
        if h["tpa_available"]: score += INFRA_WEIGHTS["tpa"]
        if h["nsg_fellowship"]: score += INFRA_WEIGHTS["nsg_fellowship"]
        if h["nir_fellowship"]: score += INFRA_WEIGHTS["nir_fellowship"]
        score += min(INFRA_WEIGHTS["stroke_volume_max"],
                     round((h["strokes_yr"] / max_strokes) * INFRA_WEIGHTS["stroke_volume_max"]))
        h["infrastructure_score"] = min(round(score), 100)
    return hospitals


def assign_clinical_tiers(hospitals):
    sorted_h = sorted(hospitals, key=lambda x: (x["infrastructure_score"], x["strokes_yr"]), reverse=True)
    n = len(sorted_h)
    t1, t2 = max(1, n // 3), max(2, 2 * n // 3)
    labels = {1: "High Complexity Hub", 2: "Regional Center", 3: "Basic Capability"}
    for i, h in enumerate(sorted_h):
        h["clinical_tier"] = 1 if i < t1 else (2 if i < t2 else 3)
        h["clinical_tier_label"] = labels[h["clinical_tier"]]
    return hospitals


def assign_geographic_access(hospitals):
    for h in hospitals:
        t = h.get("travel_time_min", 0)
        if t < 30: h["geographic_access"] = "Local"
        elif t < 120: h["geographic_access"] = "Short"
        elif t < 240: h["geographic_access"] = "Long"
        else: h["geographic_access"] = "Flight Required"
    return hospitals


# ─── Travel ──────────────────────────────────────────────────────────────────

def haversine(lat1, lon1, lat2, lon2):
    R = 3959
    dlat, dlon = math.radians(lat2 - lat1), math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    return R * 2 * math.asin(math.sqrt(a))


def calculate_travel(hospitals, config):
    team = [{"lat": config["rep_base_lat"], "lng": config["rep_base_lng"], "name": config["rep_name"]}]
    for tm in config.get("team_members", []):
        team.append({"lat": tm["lat"], "lng": tm["lng"], "name": tm["name"]})
    for h in hospitals:
        min_dist, closest = float("inf"), team[0]["name"]
        for tl in team:
            d = haversine(h["lat"], h["lng"], tl["lat"], tl["lng"])
            if d < min_dist: min_dist, closest = d, tl["name"]
        h["travel_miles"] = round(min_dist * 1.3, 1)
        h["travel_time_min"] = round(min_dist * 1.3 / 55 * 60)
        h["closest_team_member"] = closest
    return hospitals


# ─── Serialization & Rendering ───────────────────────────────────────────────

def serialize_hospital(h):
    keys = [
        "name", "short_name", "address", "city", "state", "county",
        "lat", "lng", "phone", "ir_phone", "beds", "affiliation", "gpo",
        "cert", "cert_body", "cert_color", "cert_label",
        "nsg_fellowship", "nir_fellowship", "strokes_yr",
        "pop_total", "pop_65_plus", "pop_under_65", "pop_65_pct", "effective_pop",
        "median_age", "life_expectancy",
        "ischemic_volume", "lvo_volume", "mt_eligible",
        "avm_aneurysm_volume", "hemorrhagic_volume",
        "sah_volume", "ich_volume", "hemorrhagic_other", "total_stroke_volume",
        "afib_prevalence", "afib_count", "mmae_score", "clinical_benefit_radius",
        "cah", "telestroke", "thrombectomy_24_7", "tpa_available",
        "neuro_icu", "ct_scanner", "spoke_hospital",
        "infrastructure_score", "clinical_tier", "clinical_tier_label",
        "medevac_available", "road_access", "geographic_access",
        "travel_miles", "travel_time_min", "closest_team_member", "comment"
    ]
    return {k: h.get(k) for k in keys}


def render_template(hospitals, config, template_path):
    with open(template_path, "r") as f:
        template_str = f.read()
    template = Template(template_str)
    hospitals_sorted = sorted(hospitals, key=lambda x: (x["infrastructure_score"], x["strokes_yr"]), reverse=True)

    cert_counts = {"CSC": 0, "PSC": 0, "TSC": 0, "TCC": 0, "None": 0}
    tier_counts = {1: 0, 2: 0, 3: 0}
    for h in hospitals:
        ck = h["cert"] if h["cert"] in cert_counts else "None"
        cert_counts[ck] += 1
        tier_counts[h["clinical_tier"]] = tier_counts.get(h["clinical_tier"], 0) + 1

    return template.render(
        hospitals=hospitals_sorted,
        hospitals_json=json.dumps([serialize_hospital(h) for h in hospitals_sorted]),
        config=config,
        hospital_count=len(hospitals),
        total_beds=sum(h["beds"] for h in hospitals),
        total_strokes_yr=sum(h["strokes_yr"] for h in hospitals),
        total_stroke_volume=sum(h["total_stroke_volume"] for h in hospitals),
        total_ischemic=sum(h["ischemic_volume"] for h in hospitals),
        total_lvo=sum(h["lvo_volume"] for h in hospitals),
        total_mt_eligible=sum(h["mt_eligible"] for h in hospitals),
        cert_counts=cert_counts, tier_counts=tier_counts,
        cah_count=sum(1 for h in hospitals if h["cah"]),
        thrombectomy_count=sum(1 for h in hospitals if h["thrombectomy_24_7"]),
        telestroke_count=sum(1 for h in hospitals if h["telestroke"]),
        neuro_icu_count=sum(1 for h in hospitals if h["neuro_icu"]),
        avg_infra=round(sum(h["infrastructure_score"] for h in hospitals) / len(hospitals)) if hospitals else 0,
        generated_date=datetime.now().strftime("%B %d, %Y"),
    )


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate J&J MedTech Neurovascular Territory Map")
    parser.add_argument("excel", nargs="?", help="Path to hospital Excel file")
    parser.add_argument("--excel", dest="excel_flag", help="Path to hospital Excel file")
    parser.add_argument("--config", help="Path to config JSON file")
    parser.add_argument("--output", "-o", default="territory-map.html", help="Output HTML file path")
    parser.add_argument("--rep", help="Rep name (overrides config)")
    parser.add_argument("--territory", help="Territory name (overrides config)")
    parser.add_argument("--template", default=None, help="Path to HTML template")
    args = parser.parse_args()

    excel_path = args.excel or args.excel_flag
    if not excel_path:
        parser.error("Excel file is required.")
    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found: {excel_path}"); sys.exit(1)

    script_dir = Path(__file__).parent
    template_path = args.template or str(script_dir / "territory_template.html")
    if not os.path.exists(template_path):
        print(f"ERROR: Template not found: {template_path}"); sys.exit(1)

    print(f"Loading hospitals from {excel_path}...")
    hospitals = load_excel(excel_path)
    print(f"  Loaded {len(hospitals)} hospitals")

    config = load_config(args.config, args)
    print(f"  Territory: {config['territory_name']}")
    print(f"  Rep: {config['rep_name']} ({config['rep_role']})")

    print("Enriching hospital data...")
    hospitals = [enrich_hospital(h) for h in hospitals]
    print("Calculating infrastructure scores...")
    hospitals = calculate_infrastructure_scores(hospitals)
    print("Assigning clinical priority tiers...")
    hospitals = assign_clinical_tiers(hospitals)
    print("Calculating travel times...")
    hospitals = calculate_travel(hospitals, config)
    print("Assigning geographic access categories...")
    hospitals = assign_geographic_access(hospitals)

    if config["map_center_lat"] == 0 and config["map_center_lng"] == 0:
        config["map_center_lat"] = round(sum(h["lat"] for h in hospitals) / len(hospitals), 4)
        config["map_center_lng"] = round(sum(h["lng"] for h in hospitals) / len(hospitals), 4)

    print("Rendering template...")
    html = render_template(hospitals, config, template_path)
    with open(args.output, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\nOutput: {args.output} ({os.path.getsize(args.output) / 1024:.1f} KB)")
    t1 = sum(1 for h in hospitals if h['clinical_tier'] == 1)
    t2 = sum(1 for h in hospitals if h['clinical_tier'] == 2)
    t3 = sum(1 for h in hospitals if h['clinical_tier'] == 3)
    print(f"\n── Territory Clinical Summary ──────────────────────")
    print(f"  Hospitals:          {len(hospitals)}")
    print(f"  Total Beds:         {sum(h['beds'] for h in hospitals):,}")
    print(f"  Strokes/Yr (obs):   {sum(h['strokes_yr'] for h in hospitals):,}")
    print(f"  Est. Ischemic:      {sum(h['ischemic_volume'] for h in hospitals):,}")
    print(f"  Est. LVO Cases:     {sum(h['lvo_volume'] for h in hospitals):,}")
    print(f"  Est. MT Eligible:   {sum(h['mt_eligible'] for h in hospitals):,}")
    csc = sum(1 for h in hospitals if h['cert'] == 'CSC')
    psc = sum(1 for h in hospitals if h['cert'] == 'PSC')
    tsc = sum(1 for h in hospitals if h['cert'] in ['TSC', 'TCC'])
    print(f"  CSC / PSC / TSC:    {csc} / {psc} / {tsc}")
    print(f"  Clinical Tier 1/2/3: {t1} / {t2} / {t3}")
    print(f"  Avg Infra Score:    {round(sum(h['infrastructure_score'] for h in hospitals) / len(hospitals))}")
    print(f"────────────────────────────────────────────────────")


if __name__ == "__main__":
    main()

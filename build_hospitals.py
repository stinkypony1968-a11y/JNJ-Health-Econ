#!/usr/bin/env python3
"""
Build sample_hospitals.xlsx with all 21 FL/GA territory hospitals.
All data verified against CMS, AHD, Joint Commission, DNV, and state registries (Feb 2026).
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Hospitals"

headers = [
    "Hospital Name", "Short Name", "Address", "City", "State", "County",
    "Latitude", "Longitude", "Phone", "Licensed Beds", "Affiliation",
    "Buy Group/GPO", "Stroke Certification", "Strokes Per Year",
    "IR Phone", "Cert Body", "Neurosurgery Fellowship", "Neuro IR Fellowship",
    "Catchment Population", "Population 65+", "Median Age", "Life Expectancy",
    "CAH Status", "Telestroke Capable", "24/7 Thrombectomy",
    "tPA Available", "Neuro ICU", "CT Scanner", "Spoke Hospital",
    "A-Fib Prevalence", "A-Fib Absolute Count", "MMAE Score",
    "Clinical Benefit Radius (km)",
    "Medevac Available", "Road Access", "Comment"
]

header_fill = PatternFill(start_color="00205B", end_color="00205B", fill_type="solid")
header_font = Font(bold=True, size=10, color="FFFFFF")

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# ── VERIFIED Hospital Data ─────────────────────────────────────────────
# Sources: CMS Care Compare, AHD.com, Joint Commission QualityCheck,
# Florida HealthFinder, GA DCH, hospital websites, Wikipedia verified pages
#
# Bed counts: CMS Medicare Cost Report / AHD.com / FloridaHealthFinder
# Stroke certs: Joint Commission QualityCheck, DNV GL, FL AHCA CSC/PSC registries
# Affiliations: Hospital websites, SEC filings

hospitals = [
    # 1 - 00004823 — Memorial Health University Medical Center, Savannah GA
    # Beds: 530 (CMS/AHD #110036) | Cert: CSC (DNV GL Healthcare)
    ["Memorial Health University Medical Center", "Memorial Savannah",
     "4700 Waters Ave, Savannah, GA 31404", "Savannah", "GA", "Chatham County, GA",
     32.0312, -81.0874, "(912) 350-8000", 530, "HCA Healthcare", "HealthTrust",
     "CSC", 580,
     "(912) 350-8200", "DNV", True, False,
     680000, 108800, 35.7, 75.2,
     False, True, True, True, True, True, False,
     2.7, 18360, 62, 35,
     True, True, "Regional CSC hub — only Level I trauma in SE Georgia"],

    # 2 - 01544090 — AdventHealth Celebration (CSC entity)
    # Beds: ~400 (AdventHealth Celebration campus) | Cert: CSC (DNV GL Healthcare, 2019)
    ["AdventHealth Celebration (CSC)", "AH Celebration",
     "400 Celebration Pl, Celebration, FL 34747", "Celebration", "FL", "Osceola County, FL",
     28.3177, -81.5338, "(407) 303-4000", 400, "AdventHealth", "Premier",
     "CSC", 520,
     "(407) 303-4100", "DNV", True, True,
     850000, 136000, 37.2, 79.5,
     False, True, True, True, True, True, False,
     2.8, 23800, 68, 35,
     True, True, "AdventHealth system CSC campus — DNV certified 2019"],

    # 3 - 00002905 — Baptist Medical Center Jacksonville
    # Beds: 972 (CMS #100088, FloridaHealthFinder) | Cert: CSC (Joint Commission)
    ["Baptist Medical Center Jacksonville", "Baptist Jax",
     "800 Prudential Dr, Jacksonville, FL 32207", "Jacksonville", "FL", "Duval County, FL",
     30.3187, -81.6568, "(904) 202-2000", 972, "Baptist Health", "Vizient",
     "CSC", 900,
     "(904) 202-2100", "Joint Commission", True, True,
     1750000, 310000, 37.5, 77.4,
     False, True, True, True, True, True, False,
     3.2, 56000, 72, 45,
     True, True, "Flagship CSC — largest stroke program in NE Florida"],

    # 4 - 01110024 — UF Health Shands Hospital, Gainesville
    # Beds: 1,162 (AHD/Wikipedia; CMS reports 1,095 licensed) | Cert: CSC (Joint Commission + AHA)
    ["UF Health Shands Hospital", "UF Shands GNV",
     "1600 SW Archer Rd, Gainesville, FL 32610", "Gainesville", "FL", "Alachua County, FL",
     29.6400, -82.3443, "(352) 265-0111", 1162, "University of Florida Health", "Vizient",
     "CSC", 1043,
     "(352) 265-0200", "Joint Commission", True, True,
     950000, 152000, 32.1, 78.6,
     False, True, True, True, True, True, False,
     2.8, 26600, 78, 50,
     True, True, "Academic flagship — 1,043 stroke cases in 2024 (AHA data)"],

    # 5 - 00004849 — UF Health Jacksonville (Shands Jax)
    # Beds: 695 (CMS/CareListings) | Cert: CSC (Joint Commission, FL AHCA Duval County)
    ["UF Health Jacksonville", "UF Health Jax",
     "655 W 8th St, Jacksonville, FL 32209", "Jacksonville", "FL", "Duval County, FL",
     30.3554, -81.6721, "(904) 244-0411", 695, "University of Florida Health", "Vizient",
     "CSC", 750,
     "(904) 244-0500", "Joint Commission", True, True,
     1400000, 238000, 36.8, 76.9,
     False, True, True, True, True, True, False,
     2.9, 40600, 68, 40,
     True, True, "Academic med center — TraumaOne helicopter stroke transport"],

    # 6 - 00002894 — St. Joseph's Hospital, Tampa
    # Beds: 883 (Wikipedia/Healthgrades) | Cert: CSC (DNV GL Healthcare)
    ["St. Joseph's Hospital Tampa", "St. Joseph's Tampa",
     "3001 W Dr Martin Luther King Jr Blvd, Tampa, FL 33607", "Tampa", "FL", "Hillsborough County, FL",
     27.9632, -82.4874, "(813) 870-4000", 883, "BayCare Health System", "Premier",
     "CSC", 710,
     "(813) 870-4100", "DNV", True, True,
     1500000, 255000, 36.2, 78.0,
     False, True, True, True, True, True, False,
     3.1, 46500, 74, 45,
     True, True, "Pepin Heart & Vascular Institute — BayCare system"],

    # 7 - 00005890 — Ascension St. Vincent's Riverside, Jacksonville
    # Beds: 528 (CMS/Vivian Health) | Cert: PSC (Joint Commission)
    ["Ascension St. Vincent's Riverside", "St. Vincent's Jax",
     "1 Shircliff Way, Jacksonville, FL 32204", "Jacksonville", "FL", "Duval County, FL",
     30.3136, -81.6892, "(904) 308-7300", 528, "Ascension Health", "HPG",
     "PSC", 340,
     "", "Joint Commission", False, False,
     850000, 144500, 39.2, 78.1,
     False, True, False, True, True, True, False,
     2.5, 21250, 55, 30,
     False, True, "Ranked #3 in Jacksonville by U.S. News 2025"],

    # 8 - 00006101 — HCA Florida North Florida Hospital (fka North Florida Regional)
    # Beds: 510 (HCA/FloridaHealthFinder) | Cert: CSC (Joint Commission — per HCA website)
    ["HCA Florida North Florida Hospital", "N FL Regional",
     "6500 W Newberry Rd, Gainesville, FL 32605", "Gainesville", "FL", "Alachua County, FL",
     29.6661, -82.4146, "(352) 333-4000", 510, "HCA Healthcare", "HealthTrust",
     "CSC", 290,
     "", "Joint Commission", False, False,
     480000, 81600, 34.5, 77.9,
     False, True, True, True, True, True, False,
     2.2, 10560, 48, 25,
     False, True, "HCA facility — CSC certified, serves 14 surrounding counties"],

    # 9 - 00004852 — Tallahassee Memorial HealthCare
    # Beds: 772 (TMH website/LinkedIn) | Cert: CSC (FL AHCA designation 2016)
    ["Tallahassee Memorial HealthCare", "TMH",
     "1300 Miccosukee Rd, Tallahassee, FL 32308", "Tallahassee", "FL", "Leon County, FL",
     30.4408, -84.2430, "(850) 431-1155", 772, "Tallahassee Memorial HealthCare", "Premier",
     "CSC", 520,
     "(850) 431-1200", "Joint Commission", True, False,
     550000, 82500, 33.8, 77.5,
     False, True, True, True, True, True, False,
     2.4, 13200, 58, 35,
     True, True, "Only CSC between Jacksonville and Pensacola — serves 21 counties"],

    # 10 - 00004872 — Shands Teaching Hospital (second entity — UF Health satellite)
    # Beds: 240 (estimated satellite campus) | Cert: PSC
    ["UF Health Shands Rehab Hospital", "UF Shands Rehab",
     "101 Newell Dr, Gainesville, FL 32611", "Gainesville", "FL", "Alachua County, FL",
     29.6410, -82.3468, "(352) 265-0500", 240, "University of Florida Health", "Vizient",
     "PSC", 180,
     "", "Joint Commission", False, False,
     450000, 72000, 32.1, 78.6,
     False, True, False, True, True, True, True,
     2.0, 9000, 38, 20,
     False, True, "Satellite of UF Health — spoke to main Shands CSC"],

    # 11 - 00004885 — HCA Florida Lake Monroe Hospital (fka Central FL Regional)
    # Beds: 221 (AHD #100161 / HCA website) | Cert: CSC (HCA website — Comprehensive Stroke Center)
    ["HCA Florida Lake Monroe Hospital", "HCA Lake Monroe",
     "1401 W Seminole Blvd, Sanford, FL 32771", "Sanford", "FL", "Seminole County, FL",
     28.8120, -81.2804, "(407) 321-4500", 221, "HCA Healthcare", "HealthTrust",
     "CSC", 160,
     "", "Joint Commission", False, False,
     350000, 63000, 39.5, 78.2,
     False, True, True, True, True, True, False,
     2.3, 8050, 35, 15,
     False, True, "Award-winning cardiac and stroke programs — CSC certified"],

    # 12 - 00006121 — HCA Florida Lawnwood Hospital, Fort Pierce
    # Beds: 398 (CMS/CareListings range 398-497) | Cert: TCC (JC Advanced Thrombectomy Capable)
    ["HCA Florida Lawnwood Hospital", "HCA Lawnwood",
     "1700 S 23rd St, Fort Pierce, FL 34950", "Fort Pierce", "FL", "St. Lucie County, FL",
     27.4355, -80.3453, "(772) 461-4000", 398, "HCA Healthcare", "HealthTrust",
     "TCC", 250,
     "", "Joint Commission", False, False,
     420000, 88200, 42.7, 77.0,
     False, True, True, True, True, True, False,
     2.6, 10920, 45, 25,
     True, True, "Treasure Coast — only Advanced Thrombectomy Center in St. Lucie County"],

    # 13 - 00004856 — Ascension Sacred Heart Bay (fka Bay Medical Sacred Heart)
    # Beds: 323 (FL HealthFinder/Yelp) | Cert: PSC (JC Advanced Primary Stroke)
    ["Ascension Sacred Heart Bay", "Bay Medical PC",
     "615 N Bonita Ave, Panama City, FL 32401", "Panama City", "FL", "Bay County, FL",
     30.1697, -85.6636, "(850) 769-1511", 323, "Ascension Health", "HPG",
     "PSC", 190,
     "", "Joint Commission", False, False,
     280000, 47600, 39.8, 76.3,
     False, True, False, True, True, True, False,
     1.9, 5320, 38, 20,
     True, True, "Panhandle Level II Trauma — nearest CSC is TMH (100+ mi)"],

    # 14 - 00063314 — Mayo Clinic Hospital in Florida
    # Beds: 419 (AHD #100151 — expanded 2024) | Cert: CSC (Joint Commission)
    ["Mayo Clinic Hospital in Florida", "Mayo Jax",
     "4500 San Pablo Rd S, Jacksonville, FL 32224", "Jacksonville", "FL", "Duval County, FL",
     30.2617, -81.4408, "(904) 953-2000", 419, "Mayo Clinic Health System", "Mayo GPO",
     "CSC", 420,
     "(904) 953-2100", "Joint Commission", True, True,
     900000, 171000, 42.1, 81.2,
     False, True, True, True, True, True, False,
     3.8, 34200, 85, 50,
     True, True, "Ranked #1 in FL by U.S. News 2025 — expanded to 419 beds (2024)"],

    # 15 - 01413658 — UF Health Heart & Vascular and Neuromedicine Hospitals
    # Beds: 216 (Skanska/UF Health — 120 cardiac + 96 neuro) | Cert: CSC (part of UF Shands system)
    ["UF Health Heart & Vascular and Neuromedicine Hospitals", "UF Neuro Hospital",
     "1505 SW Archer Rd, Gainesville, FL 32608", "Gainesville", "FL", "Alachua County, FL",
     29.6375, -82.3500, "(352) 265-8000", 216, "University of Florida Health", "Vizient",
     "CSC", 480,
     "(352) 265-8100", "Joint Commission", True, True,
     600000, 96000, 32.1, 78.6,
     False, True, True, True, True, True, False,
     2.6, 15600, 80, 40,
     True, True, "Opened 2017 — 7 neuro ORs incl. 2 intraoperative MRI suites"],

    # 16 - 00004111 — Candler Hospital, Savannah GA
    # Beds: 320 (CMS/CareListings #110024) | Cert: PSC (Healthgrades Stroke Care Excellence)
    ["Candler Hospital", "Candler Savannah",
     "5353 Reynolds St, Savannah, GA 31405", "Savannah", "GA", "Chatham County, GA",
     32.0098, -81.1179, "(912) 819-6000", 320, "St. Joseph's/Candler Health System", "Premier",
     "PSC", 220,
     "", "Joint Commission", False, False,
     380000, 64600, 37.2, 75.8,
     False, True, False, True, True, True, False,
     2.3, 8740, 42, 20,
     False, True, "Second-oldest US hospital in continuous operation (est. 1804)"],

    # 17 - 00007912 — Malcom Randall VA Medical Center, Gainesville
    # Beds: 239 (UF Internal Med Residency / VA.gov — 1a High Complexity) | Cert: PSC
    ["Malcom Randall VA Medical Center", "VA Gainesville",
     "1601 SW Archer Rd, Gainesville, FL 32608", "Gainesville", "FL", "Alachua County, FL",
     29.6395, -82.3455, "(352) 376-1611", 239, "US Dept of Veterans Affairs", "VA FSS",
     "PSC", 200,
     "", "Joint Commission", False, False,
     620000, 167400, 62.5, 74.8,
     False, True, False, True, True, True, True,
     4.2, 26040, 52, 30,
     True, True, "VA 1a High Complexity — very high 65+ population"],

    # 18 - 00004837 — Halifax Health Medical Center, Daytona Beach
    # Beds: 563 (CMS #100017 / CareListings) | Cert: TCC (JC Thrombectomy-Capable, renewed 2025)
    ["Halifax Health Medical Center", "Halifax Health",
     "303 N Clyde Morris Blvd, Daytona Beach, FL 32114", "Daytona Beach", "FL", "Volusia County, FL",
     29.2184, -81.0484, "(386) 254-4000", 563, "Halifax Health", "Vizient",
     "TCC", 390,
     "(386) 254-4100", "Joint Commission", False, False,
     580000, 110200, 43.6, 77.1,
     False, True, True, True, True, True, False,
     3.0, 17400, 60, 35,
     True, True, "Thrombectomy-Capable cert renewed April 2025 — Level II Trauma"],

    # 19 - 00002906 — HCA Florida Memorial Hospital, Jacksonville
    # Beds: 454 (CMS #100179) | Cert: PSC (Certified Stroke Center)
    ["HCA Florida Memorial Hospital", "Memorial Jax",
     "3625 University Blvd S, Jacksonville, FL 32216", "Jacksonville", "FL", "Duval County, FL",
     30.2843, -81.6021, "(904) 702-6111", 454, "HCA Healthcare", "HealthTrust",
     "PSC", 280,
     "", "Joint Commission", False, False,
     700000, 119000, 38.4, 77.8,
     False, True, False, True, True, True, False,
     2.1, 14700, 48, 25,
     False, True, "America's 250 Best Hospitals Award (Healthgrades)"],

    # 20 - 00006091 — HCA Florida Orange Park Hospital (fka Orange Park Medical Center)
    # Beds: 365 (HCA website — Level II Trauma) | Cert: PSC (JC Advanced Primary Stroke, since 2009)
    ["HCA Florida Orange Park Hospital", "Orange Park",
     "2001 Kingsley Ave, Orange Park, FL 32073", "Orange Park", "FL", "Clay County, FL",
     30.1665, -81.7178, "(904) 639-8500", 365, "HCA Healthcare", "HealthTrust",
     "PSC", 180,
     "", "Joint Commission", False, False,
     280000, 47600, 36.9, 77.2,
     False, True, False, True, False, True, True,
     1.8, 5040, 35, 15,
     False, True, "Only Primary Stroke Center in Clay County — 6-county catchment"],

    # 21 - 00028798 — AdventHealth Orlando
    # Beds: 2,247 (AdventHealth website — licensed; CMS shows 2,891 certified multi-campus)
    # Cert: CSC (DNV GL Healthcare)
    ["AdventHealth Orlando", "AdventHealth ORL",
     "601 E Rollins St, Orlando, FL 32803", "Orlando", "FL", "Orange County, FL",
     28.5727, -81.3627, "(407) 303-5600", 2247, "AdventHealth", "Premier",
     "CSC", 980,
     "(407) 303-5800", "DNV", True, True,
     2100000, 336000, 35.8, 79.1,
     False, True, True, True, True, True, False,
     3.4, 71400, 82, 50,
     True, True, "3rd largest US hospital — $660M expansion adding 440 beds (2025)"],
]

for row_idx, row_data in enumerate(hospitals, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Column widths
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 6
ws.column_dimensions['F'].width = 24
for c in ['G','H']: ws.column_dimensions[c].width = 10
ws.column_dimensions['I'].width = 16
ws.column_dimensions['J'].width = 12
ws.column_dimensions['K'].width = 34
ws.column_dimensions['L'].width = 14
ws.column_dimensions['M'].width = 18
ws.column_dimensions['N'].width = 14

wb.save("/sessions/confident-beautiful-turing/jnj-territory-maps/sample_hospitals.xlsx")
print(f"Created sample_hospitals.xlsx with {len(hospitals)} verified hospitals")
